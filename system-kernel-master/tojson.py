# -*- coding:utf-8 -*-
import os
import json
import datetime
import re
from frame import constant
from frame.common import start_pms_server

import openpyxl

dirname = os.path.dirname(__file__)

skip_list = ['rep.py']


# ut002037 根据给定目录获取该目录下的所有py文件
def get_case_list(dir):
    case_list = []
    for dirname, subdirs, files in os.walk(dir):
        files.sort()
        for file in files:
            if file != '__init__.py' and '.py' == os.path.splitext(file)[1]:
                file_path = os.path.join(dirname, file)
                case_list.append(file_path)
    return case_list


# ut002037 新增自动扫描excute.txt配置，并根据配置将测试文件回填到dbus文件
def autofill_dbus():
    wb = openpyxl.load_workbook('./resource/pms/dbusid.xlsx')
    ws = wb.active
    max_row = ws.max_row
    # 清空表格
    for row in range(2, max_row + 1):
        if ws.cell(row=row, column=1).value is None:
            continue
        ws.cell(row=row, column=1).value = ""
        ws.cell(row=row, column=2).value = ""

    # 获取constant.execute_file配置的测试用例
    with open(constant.execute_file, "r") as f:
        case_list = []
        lines = f.readlines()
        for line in lines:
            line = line.strip("\n")
            if not line:
                # print(line)
                continue
            if os.path.isdir(line):
                case_list.extend(get_case_list(line))
            else:
                case_list.append(line)
                continue
    # 测试用例写入dbus文件
    row = 2
    for line in case_list:
        ws.cell(row=row,
                column=2).value = line.split('/')[-1].replace(".py", "")
        # print(ws.cell(row=row, column=2).value)
        tmp = re.findall(r'\d+', ws.cell(row=row, column=2).value)
        if tmp:
            ws.cell(row=row, column=1).value = tmp[-1]
        # print(ws.cell(row=row, column=1).value)
        row = row + 1
    wb.save('./resource/pms/dbusid.xlsx')


def get_step():
    """
    自动扫描constant.execute_file配置并填入dbus表格
    """
    autofill_dbus()
    wb = openpyxl.load_workbook('./resource/pms/dbusid.xlsx')
    ws = wb.active
    max_row = ws.max_row

    info = {}

    for row in range(2, max_row + 1):
        # ut002037 新增兼容dbus表格存在空格式行问题
        # print(ws.cell(row=row, column=1).value)
        if ws.cell(row=row, column=1).value is None:
            continue
        id_ = str(ws.cell(row=row, column=1).value).strip()
        # ut002037 20210825 兼容dbus表格里的用例标题带.py和不带.py的情况
        name = str(ws.cell(row=row, column=2).value).replace(".py", "").strip()
        if id_:
            info[int(id_)] = name

    with open('./resource/pms/id2name.json', 'w') as f:
        print(info)
        json.dump(info, f)

    with open('./resource/pms/name2id.json', 'w') as f:
        info2 = {value: key for key, value in info.items()}
        json.dump(info2, f)

    # info3 = {"baseurl": "http://10.0.10.200:3000/api/v1/zbox/result",
    #          "product_id": 0,
    #          "task_id": 0,
    #          "test_type": "dbus",
    #          "user_name": "zhangsan",
    #          "testtask_name": "DDE_SP3_自动化基线用例",
    #          "area": 1}
    # with open('pms_job.json', 'w', encoding='utf8') as f:
    #     json.dump(info3, f, ensure_ascii=False)

    with open('./resource/pms/total.txt', 'a', encoding='utf8') as f:
        content = f"{datetime.datetime.now()}: {len(info)}\n"
        f.write(content)
    start_pms_server()


# if __name__ == '__main__':
#     get_step()
