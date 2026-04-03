#!/usr/bin/python3
# -*- coding: utf-8 -*-

# ****************************************************
# @Test Case ID:     001
# @Test Description: 根据提测单配置安装提测包并检测版本是否匹配
# @Test Condition:
# @Test Step:    从提测单获取安装包并自动安装对比版本
# @Test expect Result:
# @Test Remark:
# @Author:  ut002037
# *****************************************************
import os
import sys
import pytest
import logging
import subprocess
import openpyxl

from frame.common import root_exec_cmd
from frame.constant import configs_root_path

file_name = os.path.join(configs_root_path, "test.xlsx")
tsc_name = sys._getframe().f_code.co_name
debug = 1


@pytest.fixture(scope="function", autouse=True)
def env_enable_disable():
    logging.info("this is env enablle stage !")
    root_exec_cmd("cp /etc/apt/sources.list /etc/apt/sources.list.bakup", [],
                  tsc_name)
    yield
    root_exec_cmd("cp /etc/apt/sources.list.bakup /etc/apt/sources.list", [],
                  tsc_name)
    logging.info("this is env disable stage !")


# 根据新增仓库更新仓库
def update_sourefile(src_list):
    # 将转测仓库加入/etc/apt/sources.list
    cmd = 'echo ' + src_list + ' >>/etc/apt/sources.list'
    logging.info(cmd)
    root_exec_cmd(cmd, [], tsc_name)
    # 将转测源码仓库加入/etc/apt/sources.list
    src_code_list = src_list.replace("deb", "deb-src", 1)
    cmd = 'echo "' + src_code_list + '" >>/etc/apt/sources.list'
    root_exec_cmd(cmd, [], tsc_name)
    # 进行仓库更新
    if os.system("echo 1|sudo -S apt update") != 0:
        logging.error(f"update {src_list} 到 /etc/apt/sources.list failed !")
    else:
        logging.info(f"update {src_list} 到/etc/apt/sources.list succeed !")


# 根据源码包名获取源码包列表
def get_binary_deb_name(src_deb_name):
    cmd = "apt showsrc " + src_deb_name + " 2>/dev/null|grep -i Binary"
    p = subprocess.Popen(cmd,
                         shell=True,
                         stdin=subprocess.PIPE,
                         stdout=subprocess.PIPE,
                         stderr=subprocess.PIPE,
                         encoding='utf-8')
    out_msg = p.stdout.read().replace("\n", "").replace("Binary: ", "")
    err_msg = p.stderr.read().replace("\n", "")
    if err_msg:
        return err_msg
    else:
        return out_msg.split(",")


# 根据包名安装包,成功返回0
def install_specified_deb(deb_name):
    cmd = r'echo 1|sudo -S DEBIAN_FRONTEND=noninteractive apt -y -o Dpkg::Options::="--force-confnew" -o Dpkg::Options::="--force-confdef" -o Dpkg::Options::="--force-confold" install ' + deb_name
    ret = os.system(cmd)
    if ret != 0:
        logging.error(f"命令{cmd} 安装包{deb_name}失败")
    else:
        logging.info(f"命令{cmd} 安装包{deb_name}成功")
    return ret


def uninstall_specified_deb(deb_name):
    cmd = "echo 1|sudo -S apt remove " + deb_name + "--purge -y"
    cmd = r'echo 1|sudo -S DEBIAN_FRONTEND=noninteractive apt -y -o Dpkg::Options::="--force-confnew" -o Dpkg::Options::="--force-confdef" -o Dpkg::Options::="--force-confold" remove --purge ' + deb_name
    # os.system(cmd)


# 根据包名获取安装包版本
def get_deb_version(deb_name):
    cmd = "apt policy " + deb_name + " 2>/dev/null|grep -i 已安装"
    p = subprocess.Popen(cmd,
                         shell=True,
                         stdin=subprocess.PIPE,
                         stdout=subprocess.PIPE,
                         stderr=subprocess.PIPE,
                         encoding='utf-8')
    out_msg = p.stdout.read().replace("\n", "")
    err_msg = p.stderr.read().replace("\n", "")
    if err_msg:
        return err_msg
    else:
        return out_msg


# 转测单包版本与实际更新仓库后获取的包版本比较
def check_deb_version(deb_expect_version, deb_acture_version):
    if deb_expect_version in deb_acture_version:
        return 0
    else:
        return -1


def test_cve_deb_version_install_check_001():
    res = []
    # 从指定文件获取仓库表并更新仓库文件
    cfg = openpyxl.load_workbook(file_name)
    cfg_sheet = cfg['安全更新']
    max_row = cfg_sheet.max_row
    # 获取源列表
    src_lists = cfg_sheet.cell(row=1, column=1).value
    if src_lists is None:
        print("获取仓库源错误，请检查提测单表格文件！")
        pytest.skip()
    else:
        src_lists = src_lists.split("\n")

    # step1 更新源文件
    for src_list in src_lists:
        if src_list != "":
            update_sourefile(src_list)

    # step2 安装提测包并核对版本
    for row in range(3, max_row):
        # 从转测单获取源码包名
        if cfg_sheet.cell(row=row, column=1).value is None:
            print(row, "读到文件空行!!")
            break
        src_deb_name = cfg_sheet.cell(row=row,
                                      column=1).value.replace("\n", "")
        # 根据源码包名获取二进制包名列表
        bin_deb_list = get_binary_deb_name(src_deb_name)
        # 从转测单获取待安装的包的预期安装版本
        deb_expect_version = cfg_sheet.cell(row=row,
                                            column=3).value.replace("\n", "")
        # 比对安装包版本
        for deb_name in bin_deb_list:
            # 安装包，如果安装包失败记录结果并跳出此轮循环
            ret = install_specified_deb(deb_name)
            if ret != 0:
                res.append(ret)
                continue

            deb_acture_version = get_deb_version(deb_name)
            ret = check_deb_version(deb_expect_version, deb_acture_version)
            if ret:
                err_info = deb_name + " 实际版本:" + deb_acture_version + " 与预期版本:" + deb_expect_version + " 不一致"
                logging.error(err_info)
            else:
                log_info = deb_name + " 实际版本:" + deb_acture_version + " 与预期版本:" + deb_expect_version + "一致"
                logging.info(log_info)
            res.append(ret)
            # uninstall_specified_deb(deb_name)

    if res.count(0) == len(res):
        assert True
    else:
        assert False
